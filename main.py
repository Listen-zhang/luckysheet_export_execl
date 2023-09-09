import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side


def rgb_string_to_hex(rgb_string):
    # 从字符串中提取RGB分量的整数值
    try:
        if "#" in rgb_string:
            return rgb_string.replace("#", "")
        rgb = tuple(map(int, rgb_string[4:-1].split(',')))
        # 将RGB转换为十六进制并添加前缀 '#'
        hex_color = '{0:02x}{1:02x}{2:02x}'.format(rgb[0], rgb[1], rgb[2])
        return hex_color
    except:
        print('')


def map_border_style(style):
    # Map Luckysheet's border style to openpyxl's style
    border_style_map = {
        "1": "thin",
        "2": "hair",
        "3": "dotted",
        "4": "dashed",
        "5": "dashDot",
        "6": "dashDotDot",
        "7": "double",
        "8": "medium",
        "9": "mediumDashDot",
        "10": "mediumDashDotDot",
        "11": "mediumDashed",
        "12": "slantDashDot",
        "13": "thick"
        # Add other mapping relationships
    }
    return border_style_map.get(style, "thin")  # Default to 'thin' border style


def export_luckysheet_to_excel(data, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active

    for sheet_data in data:
        sheet_name = sheet_data.get("name", "Sheet")
        ws.title = sheet_name
        cell_data = sheet_data.get("celldata", [])
        config_data = sheet_data.get("config", {})
        merge_data = config_data.get("merge", {})
        column_len_data = config_data.get("columnlen", {})
        row_len_data = config_data.get("rowlen", {})
        border_info_data = config_data.get("borderInfo", [])

        for col_index, col_width in column_len_data.items():
            ws.column_dimensions[chr(65 + int(col_index))].width = col_width/10

        for row_index, row_height in row_len_data.items():
            ws.row_dimensions[int(row_index) + 1].height = row_height

        for border_info in border_info_data:
            if border_info["rangeType"] == "cell":
                row_index = border_info["value"]["row_index"]
                col_index = border_info["value"]["col_index"]
                cell = ws.cell(row=row_index + 1, column=col_index + 1)
                apply_border_styles(cell, border_info["value"])
            elif border_info["rangeType"] == "range":
                for range_data in border_info["range"]:
                    apply_border_styles_range(ws, range_data, border_info["color"], border_info["style"])

        for cell_info in cell_data:
            row_index = cell_info["r"]
            col_index = cell_info["c"]
            try:
                cell_value = cell_info["v"]["v"]
            except:
                cell_value = None
                print('')
            cell = ws.cell(row=row_index + 1, column=col_index + 1, value=cell_value)

            font_info = cell_info["v"]
            apply_font_styles(cell, font_info)

        for merge_key, merge_info in merge_data.items():
            row_index = merge_info["r"]
            col_index = merge_info["c"]
            row_span = merge_info["rs"]
            col_span = merge_info["cs"]
            merge_range(ws, row_index, col_index, row_span, col_span)
    wb.save(output_file)

def font_type(f_id):
    font_type_map = {
        "0": "Times New Roman",
        "1": "Arial",
        "2": "Verdana",
        "3": "微软雅黑",
        "4": "宋体",
        "5": "黑体",
        "6": "楷体",
        "7": "仿宋",
        "8": "新宋体",
        "9": "华文新魏",
        "10": "华文行楷",
        "11": "华文隶书",
        "12": "thick"
        # Add other mapping relationships
    }
    return font_type_map[str(f_id)]

def apply_font_styles(cell, font_info):
    font = Font()
    if font_info.get("fs"):
        font.size = font_info["fs"]
    if font_info.get("bl") and font_info["bl"] == 1:
        font.bold = True
    if font_info.get("it") and font_info["it"] == 1:
        font.italic = True
    if font_info.get("fc"):
        font.color = rgb_string_to_hex(font_info["fc"])
    if font_info.get("ff"):
        font.name = font_type(font_info["ff"])
    if font_info.get("bg"):
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor=rgb_string_to_hex(font_info["bg"]))

    cell.font = font


def apply_border_styles(cell, border_info):
    border = Border()
    if border_info.get("l"):
        border.left = Side(style=map_border_style(border_info["l"]["style"]),
                           color=rgb_string_to_hex(border_info["l"]["color"]))
    if border_info.get("r"):
        border.right = Side(style=map_border_style(border_info["r"]["style"]),
                            color=rgb_string_to_hex(border_info["r"]["color"]))
    if border_info.get("t"):
        border.top = Side(style=map_border_style(border_info["t"]["style"]),
                          color=rgb_string_to_hex(border_info["t"]["color"]))
    if border_info.get("b"):
        border.bottom = Side(style=map_border_style(border_info["b"]["style"]),
                             color=rgb_string_to_hex(border_info["b"]["color"]))

    cell.border = border


def apply_border_styles_range(ws, range_data, color, style):
    rows = range_data.get("row", [])
    cols = range_data.get("column", [])
    for row_index in range(rows[0], rows[1] + 1):
        for col_index in range(cols[0], cols[1] + 1):
            cell = ws.cell(row=row_index + 1, column=col_index + 1)
            border = Border()
            border.left = Side(style=map_border_style(style), color=color.replace('#', ''))
            border.right = Side(style=map_border_style(style), color=color.replace('#', ''))
            border.top = Side(style=map_border_style(style), color=color.replace('#', ''))
            border.bottom = Side(style=map_border_style(style), color=color.replace('#', ''))
            cell.border = border


def merge_range(ws, row_index, col_index, row_span, col_span):
    ws.merge_cells(start_row=row_index + 1, start_column=col_index + 1, end_row=row_index + row_span, end_column=col_index + col_span)

if __name__ == "__main__":
    excel_data = [{
        "name": "Cell",
        "config": {
            "merge": {
                "13_5": {
                    "r": 13,
                    "c": 5,
                    "rs": 3,
                    "cs": 1
                },
                "13_7": {
                    "r": 13,
                    "c": 7,
                    "rs": 3,
                    "cs": 2
                },
                "14_2": {
                    "r": 14,
                    "c": 2,
                    "rs": 1,
                    "cs": 2
                },
                "15_10": {
                    "r": 15,
                    "c": 10,
                    "rs": 4,
                    "cs": 3
                }
            },
            "borderInfo": [
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 3,
                        "col_index": 3,
                        "l": {
                            "style": 10,
                            "color": "rgb(255, 0, 0)"
                        },
                        "r": {
                            "style": 10,
                            "color": "rgb(255, 0, 0)"
                        },
                        "t": {
                            "style": 10,
                            "color": "rgb(255, 0, 0)"
                        },
                        "b": {
                            "style": 10,
                            "color": "rgb(255, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 3,
                        "col_index": 4,
                        "l": {
                            "style": 10,
                            "color": "rgb(255, 0, 0)"
                        },
                        "r": {
                            "style": 10,
                            "color": "rgb(255, 0, 0)"
                        },
                        "t": {
                            "style": 10,
                            "color": "rgb(255, 0, 0)"
                        },
                        "b": {
                            "style": 10,
                            "color": "rgb(255, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 3,
                        "col_index": 5,
                        "l": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "r": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "t": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "b": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 3,
                        "col_index": 6,
                        "l": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "r": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "t": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "b": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 3,
                        "col_index": 7,
                        "l": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "r": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "t": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "b": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 3,
                        "col_index": 8,
                        "l": {
                            "style": 1,
                            "color": "rgb(255, 0, 0)"
                        },
                        "r": {
                            "style": 1,
                            "color": "rgb(255, 0, 0)"
                        },
                        "t": {
                            "style": 1,
                            "color": "rgb(255, 0, 0)"
                        },
                        "b": {
                            "style": 1,
                            "color": "rgb(255, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 5,
                        "col_index": 2,
                        "l": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "r": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "t": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "b": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 5,
                        "col_index": 3,
                        "l": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "r": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "t": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "b": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 5,
                        "col_index": 4,
                        "l": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "r": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "t": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "b": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 5,
                        "col_index": 5,
                        "l": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "r": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "t": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        },
                        "b": {
                            "style": 9,
                            "color": "rgb(255, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 5,
                        "col_index": 6,
                        "l": {
                            "style": 1,
                            "color": "rgb(255, 0, 0)"
                        },
                        "r": {
                            "style": 1,
                            "color": "rgb(255, 0, 0)"
                        },
                        "t": {
                            "style": 1,
                            "color": "rgb(255, 0, 0)"
                        },
                        "b": {
                            "style": 1,
                            "color": "rgb(255, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 5,
                        "col_index": 7,
                        "l": {
                            "style": 1,
                            "color": "rgb(255, 0, 0)"
                        },
                        "r": {
                            "style": 1,
                            "color": "rgb(255, 0, 0)"
                        },
                        "t": {
                            "style": 1,
                            "color": "rgb(255, 0, 0)"
                        },
                        "b": {
                            "style": 1,
                            "color": "rgb(255, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 5,
                        "col_index": 8,
                        "l": {
                            "style": 2,
                            "color": "rgb(255, 0, 0)"
                        },
                        "r": {
                            "style": 2,
                            "color": "rgb(255, 0, 0)"
                        },
                        "t": {
                            "style": 2,
                            "color": "rgb(255, 0, 0)"
                        },
                        "b": {
                            "style": 2,
                            "color": "rgb(255, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 7,
                        "col_index": 2,
                        "l": {
                            "style": 9,
                            "color": "rgb(0, 0, 255)"
                        },
                        "r": {
                            "style": 9,
                            "color": "rgb(0, 0, 255)"
                        },
                        "t": {
                            "style": 9,
                            "color": "rgb(0, 0, 255)"
                        },
                        "b": {
                            "style": 9,
                            "color": "rgb(0, 0, 255)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 7,
                        "col_index": 3,
                        "l": {
                            "style": 9,
                            "color": "rgb(0, 0, 255)"
                        },
                        "r": {
                            "style": 9,
                            "color": "rgb(0, 0, 255)"
                        },
                        "t": {
                            "style": 9,
                            "color": "rgb(0, 0, 255)"
                        },
                        "b": {
                            "style": 9,
                            "color": "rgb(0, 0, 255)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 7,
                        "col_index": 5,
                        "l": {
                            "style": 2,
                            "color": "rgb(154, 205, 50)"
                        },
                        "t": {
                            "style": 2,
                            "color": "rgb(154, 205, 50)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 7,
                        "col_index": 6,
                        "r": {
                            "style": 2,
                            "color": "rgb(154, 205, 50)"
                        },
                        "t": {
                            "style": 2,
                            "color": "rgb(154, 205, 50)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 7,
                        "col_index": 8,
                        "r": {
                            "style": 9,
                            "color": "rgb(0, 0, 0)"
                        },
                        "b": {
                            "style": 9,
                            "color": "rgb(0, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 7,
                        "col_index": 9,
                        "l": {
                            "style": 9,
                            "color": "rgb(0, 0, 0)"
                        },
                        "b": {
                            "style": 9,
                            "color": "rgb(0, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 8,
                        "col_index": 2,
                        "l": {
                            "style": 9,
                            "color": "rgb(0, 0, 255)"
                        },
                        "r": {
                            "style": 9,
                            "color": "rgb(0, 0, 255)"
                        },
                        "t": {
                            "style": 9,
                            "color": "rgb(0, 0, 255)"
                        },
                        "b": {
                            "style": 9,
                            "color": "rgb(0, 0, 255)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 8,
                        "col_index": 3,
                        "l": {
                            "style": 9,
                            "color": "rgb(0, 0, 255)"
                        },
                        "r": {
                            "style": 9,
                            "color": "rgb(0, 0, 255)"
                        },
                        "t": {
                            "style": 9,
                            "color": "rgb(0, 0, 255)"
                        },
                        "b": {
                            "style": 9,
                            "color": "rgb(0, 0, 255)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 8,
                        "col_index": 5,
                        "l": {
                            "style": 2,
                            "color": "rgb(154, 205, 50)"
                        },
                        "b": {
                            "style": 2,
                            "color": "rgb(154, 205, 50)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 8,
                        "col_index": 6,
                        "r": {
                            "style": 2,
                            "color": "rgb(154, 205, 50)"
                        },
                        "b": {
                            "style": 2,
                            "color": "rgb(154, 205, 50)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 8,
                        "col_index": 8,
                        "r": {
                            "style": 9,
                            "color": "rgb(0, 0, 0)"
                        },
                        "t": {
                            "style": 9,
                            "color": "rgb(0, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 8,
                        "col_index": 9,
                        "l": {
                            "style": 9,
                            "color": "rgb(0, 0, 0)"
                        },
                        "t": {
                            "style": 9,
                            "color": "rgb(0, 0, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 10,
                        "col_index": 2,
                        "l": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        },
                        "t": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        },
                        "b": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 10,
                        "col_index": 3,
                        "r": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        },
                        "t": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        },
                        "b": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 10,
                        "col_index": 5,
                        "l": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        },
                        "r": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        },
                        "t": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 10,
                        "col_index": 6,
                        "l": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        },
                        "r": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        },
                        "t": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 10,
                        "col_index": 7,
                        "l": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        },
                        "r": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        },
                        "t": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 11,
                        "col_index": 2,
                        "l": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        },
                        "t": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        },
                        "b": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 11,
                        "col_index": 3,
                        "r": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        },
                        "t": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        },
                        "b": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 11,
                        "col_index": 5,
                        "l": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        },
                        "r": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 11,
                        "col_index": 6,
                        "l": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        },
                        "r": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 11,
                        "col_index": 7,
                        "l": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        },
                        "r": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 12,
                        "col_index": 2,
                        "l": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        },
                        "t": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        },
                        "b": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 12,
                        "col_index": 3,
                        "r": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        },
                        "t": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        },
                        "b": {
                            "style": 1,
                            "color": "rgb(144, 238, 144)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 12,
                        "col_index": 5,
                        "l": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        },
                        "r": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        },
                        "b": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 12,
                        "col_index": 6,
                        "l": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        },
                        "r": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        },
                        "b": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 12,
                        "col_index": 7,
                        "l": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        },
                        "r": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        },
                        "b": {
                            "style": 1,
                            "color": "rgb(205, 205, 0)"
                        }
                    }
                },
                {
                    "rangeType": "range",
                    "borderType": "border-none",
                    "style": "2",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                3,
                                3
                            ],
                            "column": [
                                3,
                                4
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-all",
                    "style": "2",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                3,
                                3
                            ],
                            "column": [
                                3,
                                4
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-none",
                    "style": "4",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                3,
                                3
                            ],
                            "column": [
                                4,
                                4
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-all",
                    "style": "3",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                3,
                                3
                            ],
                            "column": [
                                4,
                                4
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-none",
                    "style": "3",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                3,
                                3
                            ],
                            "column": [
                                3,
                                3
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-all",
                    "style": "2",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                3,
                                3
                            ],
                            "column": [
                                3,
                                3
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-none",
                    "style": "2",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                3,
                                3
                            ],
                            "column": [
                                5,
                                5
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-none",
                    "style": "2",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                3,
                                3
                            ],
                            "column": [
                                8,
                                8
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-none",
                    "style": "2",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                3,
                                3
                            ],
                            "column": [
                                3,
                                8
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-all",
                    "style": "4",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                3,
                                3
                            ],
                            "column": [
                                7,
                                7
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-all",
                    "style": "1",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                3,
                                3
                            ],
                            "column": [
                                8,
                                8
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-all",
                    "style": "5",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                3,
                                3
                            ],
                            "column": [
                                6,
                                6
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-all",
                    "style": "6",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                3,
                                3
                            ],
                            "column": [
                                5,
                                5
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-none",
                    "style": "3",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                3,
                                3
                            ],
                            "column": [
                                4,
                                4
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-all",
                    "style": "3",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                3,
                                3
                            ],
                            "column": [
                                4,
                                4
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-all",
                    "style": "2",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                3,
                                3
                            ],
                            "column": [
                                3,
                                3
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-none",
                    "style": "2",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                5,
                                5
                            ],
                            "column": [
                                2,
                                9
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-all",
                    "style": "9",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                5,
                                5
                            ],
                            "column": [
                                5,
                                5
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-all",
                    "style": "8",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                5,
                                5
                            ],
                            "column": [
                                6,
                                6
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-all",
                    "style": "13",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                5,
                                5
                            ],
                            "column": [
                                7,
                                7
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-all",
                    "style": "13",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                6,
                                6
                            ],
                            "column": [
                                11,
                                11
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-all",
                    "style": "10",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                5,
                                5
                            ],
                            "column": [
                                4,
                                4
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-all",
                    "style": "11",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                5,
                                5
                            ],
                            "column": [
                                3,
                                3
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-none",
                    "style": "11",
                    "color": "#ff0000",
                    "range": [
                        {
                            "row": [
                                7,
                                8
                            ],
                            "column": [
                                2,
                                3
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-all",
                    "style": "3",
                    "color": "#0000ff",
                    "range": [
                        {
                            "row": [
                                7,
                                8
                            ],
                            "column": [
                                2,
                                3
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-inside",
                    "style": "3",
                    "color": "#0000ff",
                    "range": [
                        {
                            "row": [
                                7,
                                8
                            ],
                            "column": [
                                8,
                                9
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-inside",
                    "style": "9",
                    "color": "#0000ff",
                    "range": [
                        {
                            "row": [
                                7,
                                8
                            ],
                            "column": [
                                8,
                                9
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-inside",
                    "style": "2",
                    "color": "#0000ff",
                    "range": [
                        {
                            "row": [
                                7,
                                8
                            ],
                            "column": [
                                8,
                                9
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-inside",
                    "style": "9",
                    "color": "#0000ff",
                    "range": [
                        {
                            "row": [
                                7,
                                8
                            ],
                            "column": [
                                8,
                                9
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 10,
                        "col_index": 10,
                        "l": {
                            "color": "#ff0000",
                            "style": "13"
                        },
                        "r": {
                            "color": "#ff0000",
                            "style": "13"
                        },
                        "t": {
                            "color": "#ff0000",
                            "style": "13"
                        },
                        "b": {
                            "color": "#ff0000",
                            "style": "13"
                        }
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 6,
                        "col_index": 11,
                        "l": None,
                        "r": None,
                        "t": None,
                        "b": None
                    }
                },
                {
                    "rangeType": "cell",
                    "value": {
                        "row_index": 10,
                        "col_index": 10,
                        "l": None,
                        "r": None,
                        "t": None,
                        "b": None
                    }
                },
                {
                    "rangeType": "range",
                    "borderType": "border-outside",
                    "style": "13",
                    "color": "#00ff00",
                    "range": [
                        {
                            "row": [
                                10,
                                12
                            ],
                            "column": [
                                2,
                                3
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-horizontal",
                    "style": "13",
                    "color": "#00ff00",
                    "range": [
                        {
                            "row": [
                                10,
                                12
                            ],
                            "column": [
                                2,
                                3
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-vertical",
                    "style": "13",
                    "color": "#ff9900",
                    "range": [
                        {
                            "row": [
                                10,
                                12
                            ],
                            "column": [
                                5,
                                7
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-outside",
                    "style": "13",
                    "color": "#ff9900",
                    "range": [
                        {
                            "row": [
                                10,
                                12
                            ],
                            "column": [
                                5,
                                7
                            ]
                        }
                    ]
                },
                {
                    "rangeType": "range",
                    "borderType": "border-none",
                    "style": "1",
                    "color": "#ff9900",
                    "range": [
                        {
                            "row": [
                                19,
                                19
                            ],
                            "column": [
                                6,
                                6
                            ]
                        }
                    ]
                }
            ],
            "rowlen": {
                "0": 20,
                "1": 20,
                "2": 20,
                "3": 20,
                "4": 20,
                "5": 20,
                "6": 20,
                "7": 20,
                "8": 20,
                "9": 20,
                "10": 20,
                "11": 20,
                "12": 20,
                "13": 20,
                "14": 20,
                "15": 20,
                "16": 20,
                "17": 31,
                "18": 20,
                "19": 20,
                "20": 20,
                "21": 20,
                "22": 20,
                "23": 20,
                "24": 20,
                "25": 79,
                "26": 20,
                "27": 20,
                "28": 80,
                "29": 36
            },
            "columnlen": {
                "0": 131,
                "2": 153,
                "3": 128,
                "4": 136,
                "5": 122,
                "6": 138,
                "7": 131,
                "8": 128,
                "9": 140,
                "10": 144
            },
            "rowhidden": {
                "30": 0,
                "31": 0
            },
            "customHeight": {
                "29": 1
            },
            "customWidth": {
                "2": 1
            }
        },
        "index": "0",
        "zoomRatio": 1,
        "order": "0",
        "column": 18,
        "row": 36,
        "status": 1,
        "celldata": [{"r": 0, "c": 0, "v": {"customKey": {"a": 1}, "bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11,
                                            "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1, "v": 1,
                                            "ct": {"fa": "General", "t": "n"}, "m": "1"}}, {"r": 0, "c": 1,
                                                                                            "v": {"bg": None, "bl": 0,
                                                                                                  "it": 0, "ff": 0,
                                                                                                  "fs": 11,
                                                                                                  "fc": "rgb(51, 51, 51)",
                                                                                                  "ht": 1, "vt": 1,
                                                                                                  "v": 2, "ct": {
                                                                                                    "fa": "General",
                                                                                                    "t": "n"},
                                                                                                  "m": "2"}},
                     {"r": 0, "c": 2,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1,
                            "v": 3, "ct": {"fa": "General", "t": "n"}, "m": "3"}}, {"r": 0, "c": 3,
                                                                                    "v": {"bg": None, "bl": 0, "it": 0,
                                                                                          "ff": 0, "fs": 11,
                                                                                          "fc": "rgb(51, 51, 51)",
                                                                                          "ht": 1, "vt": 1, "v": 0,
                                                                                          "ct": {"fa": "General",
                                                                                                 "t": "n"}, "m": "0",
                                                                                          "f": "=Formula!D3+Formula!D4"}},
                     {"r": 0, "c": 4,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 0, "c": 5,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 0, "c": 6,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 0, "c": 7,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 0, "c": 8,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 0, "c": 9,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 0, "c": 10,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 1, "c": 0,
                                                                                     "v": {"v": "Background",
                                                                                           "ct": {"fa": "General",
                                                                                                  "t": "g"},
                                                                                           "m": "Background",
                                                                                           "bg": None, "bl": 1, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 1, "c": 1,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}},
                     {"r": 1, "c": 2,
                      "v": {"bg": "rgb(30, 144, 255)", "bl": 0, "it": 0, "ff": 0, "fs": 11,
                            "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1}}, {"r": 1, "c": 3,
                                                                          "v": {"bg": None, "bl": 0,
                                                                                "it": 0, "ff": 0,
                                                                                "fs": 11,
                                                                                "fc": "rgb(51, 51, 51)",
                                                                                "ht": 1, "vt": 1}},
                     {"r": 1, "c": 4,
                      "v": {"bg": "rgb(0, 255, 0)", "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                            "ht": 1, "vt": 1}}, {"r": 1, "c": 5, "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11,
                                                                       "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1}},
                     {"r": 1, "c": 6,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 1, "c": 7,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 1, "c": 8,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 1, "c": 9,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 1, "c": 10,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 2, "c": 0,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 2, "c": 1,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 2, "c": 2,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 2, "c": 3,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 2, "c": 4,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 2, "c": 5,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 2, "c": 6,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 2, "c": 7,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 2, "c": 8,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 2, "c": 9,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 2, "c": 10,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 3, "c": 0,
                      "v": {"v": "Border", "ct": {"fa": "General", "t": "g"}, "m": "Border", "bg": None, "bl": 1,
                            "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1}}, {"r": 3, "c": 1,
                                                                                                      "v": {"bg": None,
                                                                                                            "bl": 0,
                                                                                                            "it": 0,
                                                                                                            "ff": 0,
                                                                                                            "fs": 11,
                                                                                                            "fc": "rgb(51, 51, 51)",
                                                                                                            "ht": 1,
                                                                                                            "vt": 1}},
                     {"r": 3, "c": 2,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 3, "c": 3,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 3, "c": 4,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 3, "c": 5,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 3, "c": 6,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 3, "c": 7,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 3, "c": 8,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 3, "c": 9,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 3, "c": 10,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 4, "c": 0,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 4, "c": 1,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 4, "c": 2,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 4, "c": 3,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 4, "c": 4,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 4, "c": 5,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 4, "c": 6,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 4, "c": 7,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 4, "c": 8,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 4, "c": 9,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 4, "c": 10,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 5, "c": 0,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1,
                            "ct": {"fa": "General", "t": "inlineStr", "s": [
                                {"ff": "Arial", "fc": "rgb(255, 0, 0)", "fs": 12, "cl": 0, "un": 0, "bl": 0, "it": 0,
                                 "v": "Inline"},
                                {"ff": "Arial", "fc": "#000000", "fs": 12, "cl": 0, "un": 0, "bl": 0, "it": 0,
                                 "v": " "},
                                {"ff": "Arial", "fc": "#000000", "fs": 16, "cl": 1, "un": 0, "bl": 0, "it": 1,
                                 "v": "Style"},
                                {"ff": "Arial", "fc": "#000000", "fs": 12, "cl": 0, "un": 0, "bl": 0, "it": 0,
                                 "v": " "},
                                {"ff": "Arial", "fc": "#000000", "fs": 12, "cl": 0, "un": 0, "bl": 1, "it": 0,
                                 "v": "Cell"}]}}}, {"r": 5, "c": 1,
                                                    "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11,
                                                          "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1}}, {"r": 5, "c": 2,
                                                                                                        "v": {
                                                                                                            "bg": None,
                                                                                                            "bl": 0,
                                                                                                            "it": 0,
                                                                                                            "ff": 0,
                                                                                                            "fs": 11,
                                                                                                            "fc": "rgb(51, 51, 51)",
                                                                                                            "ht": 1,
                                                                                                            "vt": 1}},
                     {"r": 5, "c": 3,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 5, "c": 4,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 5, "c": 5,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 5, "c": 6,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 5, "c": 7,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 5, "c": 8,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 5, "c": 9,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 5, "c": 10,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 6, "c": 0,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 6, "c": 1,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 6, "c": 2,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 6, "c": 3,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 6, "c": 4,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 6, "c": 5,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 6, "c": 6,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 6, "c": 7,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 6, "c": 8,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 6, "c": 9,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 6, "c": 10,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 7, "c": 0,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 7, "c": 1,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 7, "c": 2,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 7, "c": 3,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 7, "c": 4,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 7, "c": 5,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 7, "c": 6,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 7, "c": 7,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 7, "c": 8,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 7, "c": 9,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 7, "c": 10,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 8, "c": 0,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 8, "c": 1,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 8, "c": 2,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 8, "c": 3,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 8, "c": 4,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 8, "c": 5,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 8, "c": 6,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 8, "c": 7,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 8, "c": 8,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 8, "c": 9,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 8, "c": 10,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 9, "c": 0,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 9, "c": 1,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 9, "c": 2,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 9, "c": 3,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 9, "c": 4,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 9, "c": 5,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 9, "c": 6,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 9, "c": 7,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 9, "c": 8,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 9, "c": 9,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 9, "c": 10,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 10, "c": 0,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 10, "c": 1,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 10, "c": 2,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 10, "c": 3,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 10, "c": 4,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 10, "c": 5,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 10, "c": 6,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 10, "c": 7,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 10, "c": 8,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 10, "c": 9,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 10, "c": 10,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 11, "c": 0,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 11, "c": 1,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 11, "c": 2,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 11, "c": 3,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 11, "c": 4,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 11, "c": 5,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 11, "c": 6,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 11, "c": 7,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 11, "c": 8,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 11, "c": 9,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 11, "c": 10,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 12, "c": 0,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 12, "c": 1,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 12, "c": 2,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 12, "c": 3,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 12, "c": 4,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 12, "c": 5,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 12, "c": 6,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 12, "c": 7,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 12, "c": 8,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 12, "c": 9,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 12, "c": 10,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 13, "c": 0,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 13, "c": 1,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 13, "c": 2,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 13, "c": 3,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 13, "c": 4,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 13, "c": 5,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1,
                                                                         "mc": {"r": 13, "c": 5, "rs": 3, "cs": 1}}},
                     {"r": 13, "c": 6,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 13, "c": 7,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1, "mc": {"r": 13, "c": 7, "rs": 3, "cs": 2}}},
                     {"r": 13, "c": 8, "v": {"mc": {"r": 13, "c": 7}}}, {"r": 13, "c": 9,
                                                                         "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                               "fs": 11, "fc": "rgb(51, 51, 51)",
                                                                               "ht": 1, "vt": 1}}, {"r": 13, "c": 10,
                                                                                                    "v": {"bg": None,
                                                                                                          "bl": 0,
                                                                                                          "it": 0,
                                                                                                          "ff": 0,
                                                                                                          "fs": 11,
                                                                                                          "fc": "rgb(51, 51, 51)",
                                                                                                          "ht": 1,
                                                                                                          "vt": 1}},
                     {"r": 14, "c": 0,
                      "v": {"v": "Span", "ct": {"fa": "General", "t": "g"}, "m": "Span", "bg": None, "bl": 1, "it": 0,
                            "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1}}, {"r": 14, "c": 1,
                                                                                             "v": {"bg": None, "bl": 0,
                                                                                                   "it": 0, "ff": 0,
                                                                                                   "fs": 11,
                                                                                                   "fc": "rgb(51, 51, 51)",
                                                                                                   "ht": 1, "vt": 1}},
                     {"r": 14, "c": 2,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1,
                            "mc": {"r": 14, "c": 2, "rs": 1, "cs": 2}}},
                     {"r": 14, "c": 3, "v": {"mc": {"r": 14, "c": 2}}}, {"r": 14, "c": 4,
                                                                         "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                               "fs": 11, "fc": "rgb(51, 51, 51)",
                                                                               "ht": 1, "vt": 1}},
                     {"r": 14, "c": 5, "v": {"mc": {"r": 13, "c": 5}}}, {"r": 14, "c": 6,
                                                                         "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                               "fs": 11, "fc": "rgb(51, 51, 51)",
                                                                               "ht": 1, "vt": 1}},
                     {"r": 14, "c": 7, "v": {"mc": {"r": 13, "c": 7}}},
                     {"r": 14, "c": 8, "v": {"mc": {"r": 13, "c": 7}}}, {"r": 14, "c": 9,
                                                                         "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                               "fs": 11, "fc": "rgb(51, 51, 51)",
                                                                               "ht": 1, "vt": 1}}, {"r": 14, "c": 10,
                                                                                                    "v": {"bg": None,
                                                                                                          "bl": 0,
                                                                                                          "it": 0,
                                                                                                          "ff": 0,
                                                                                                          "fs": 11,
                                                                                                          "fc": "rgb(51, 51, 51)",
                                                                                                          "ht": 1,
                                                                                                          "vt": 1}},
                     {"r": 15, "c": 0,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 15, "c": 1,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 15, "c": 2,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 15, "c": 3,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 15, "c": 4,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 15, "c": 5, "v": {"mc": {"r": 13, "c": 5}}}, {"r": 15, "c": 6,
                                                                                           "v": {"bg": None, "bl": 0,
                                                                                                 "it": 0, "ff": 0,
                                                                                                 "fs": 11,
                                                                                                 "fc": "rgb(51, 51, 51)",
                                                                                                 "ht": 1, "vt": 1}},
                     {"r": 15, "c": 7, "v": {"mc": {"r": 13, "c": 7}}},
                     {"r": 15, "c": 8, "v": {"mc": {"r": 13, "c": 7}}}, {"r": 15, "c": 9,
                                                                         "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                               "fs": 11, "fc": "rgb(51, 51, 51)",
                                                                               "ht": 1, "vt": 1}}, {"r": 15, "c": 10,
                                                                                                    "v": {"bg": None,
                                                                                                          "bl": 0,
                                                                                                          "it": 0,
                                                                                                          "ff": 0,
                                                                                                          "fs": 11,
                                                                                                          "fc": "rgb(51, 51, 51)",
                                                                                                          "ht": 1,
                                                                                                          "vt": 1,
                                                                                                          "mc": {
                                                                                                              "r": 15,
                                                                                                              "c": 10,
                                                                                                              "rs": 4,
                                                                                                              "cs": 3}}},
                     {"r": 15, "c": 11, "v": {"mc": {"r": 15, "c": 10}}},
                     {"r": 15, "c": 12, "v": {"mc": {"r": 15, "c": 10}}}, {"r": 16, "c": 0,
                                                                           "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                                 "fs": 11, "fc": "rgb(51, 51, 51)",
                                                                                 "ht": 1, "vt": 1}}, {"r": 16, "c": 1,
                                                                                                      "v": {"bg": None,
                                                                                                            "bl": 0,
                                                                                                            "it": 0,
                                                                                                            "ff": 0,
                                                                                                            "fs": 11,
                                                                                                            "fc": "rgb(51, 51, 51)",
                                                                                                            "ht": 1,
                                                                                                            "vt": 1}},
                     {"r": 16, "c": 2,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 16, "c": 3,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 16, "c": 4,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 16, "c": 5,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 16, "c": 6,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 16, "c": 7,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 16, "c": 8,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 16, "c": 9,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 16, "c": 10,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1,
                            "mc": {"r": 15, "c": 10}}}, {"r": 16, "c": 11, "v": {"mc": {"r": 15, "c": 10}}},
                     {"r": 16, "c": 12, "v": {"mc": {"r": 15, "c": 10}}}, {"r": 17, "c": 0, "v": {"v": "Font", "ct": {
                "fa": "General", "t": "g"}, "m": "Font", "bg": None, "bl": 1, "it": 0, "ff": 0, "fs": 11,
                                                                                                  "fc": "rgb(51, 51, 51)",
                                                                                                  "ht": 1, "vt": 1}},
                     {"r": 17, "c": 1,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 17, "c": 2,
                                        "v": {"v": "Luckysheet", "ct": {"fa": "General", "t": "g"}, "bg": None, "bl": 0,
                                              "it": 0, "ff": 0, "fs": "11", "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1,
                                              "m": "Luckysheet"}}, {"r": 17, "c": 3, "v": {"v": "Luckysheet",
                                                                                           "ct": {"fa": "General",
                                                                                                  "t": "g"}, "bg": None,
                                                                                           "bl": 0, "it": 0, "ff": 0,
                                                                                           "fs": 13,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1,
                                                                                           "m": "Luckysheet"}},
                     {"r": 17, "c": 4,
                      "v": {"v": "Luckysheet", "ct": {"fa": "General", "t": "g"}, "bg": None, "bl": 0, "it": 0, "ff": 9,
                            "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1, "m": "Luckysheet"}}, {"r": 17, "c": 5,
                                                                                                       "v": {
                                                                                                           "v": "Luckysheet",
                                                                                                           "ct": {
                                                                                                               "fa": "General",
                                                                                                               "t": "g"},
                                                                                                           "bg": None,
                                                                                                           "bl": 0,
                                                                                                           "it": 0,
                                                                                                           "ff": 0,
                                                                                                           "fs": 13,
                                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                                           "ht": 1,
                                                                                                           "vt": 1,
                                                                                                           "m": "Luckysheet"}},
                     {"r": 17, "c": 6,
                      "v": {"v": "Luckysheet", "ct": {"fa": "General", "t": "g"}, "bg": "rgb(255, 215, 0)", "bl": 0,
                            "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1, "m": "Luckysheet"}},
                     {"r": 17, "c": 7,
                      "v": {"v": "Luckysheet", "ct": {"fa": "General", "t": "g"}, "bg": None, "bl": 0, "it": 0, "ff": 0,
                            "fs": 11, "fc": "rgb(67, 110, 238)", "ht": 1, "vt": 1, "m": "Luckysheet"}},
                     {"r": 17, "c": 8,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 17, "c": 9,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 17, "c": 10,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1, "mc": {"r": 15, "c": 10}}},
                     {"r": 17, "c": 11, "v": {"mc": {"r": 15, "c": 10}}},
                     {"r": 17, "c": 12, "v": {"mc": {"r": 15, "c": 10}}}, {"r": 18, "c": 0,
                                                                           "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                                 "fs": 11, "fc": "rgb(51, 51, 51)",
                                                                                 "ht": 1, "vt": 1}}, {"r": 18, "c": 1,
                                                                                                      "v": {"bg": None,
                                                                                                            "bl": 0,
                                                                                                            "it": 0,
                                                                                                            "ff": 0,
                                                                                                            "fs": 11,
                                                                                                            "fc": "rgb(51, 51, 51)",
                                                                                                            "ht": 1,
                                                                                                            "vt": 1}},
                     {"r": 18, "c": 2,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 18, "c": 3,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 18, "c": 4,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 18, "c": 5,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 18, "c": 6,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 18, "c": 7,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 18, "c": 8,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 18, "c": 9,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 18, "c": 10,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1,
                            "mc": {"r": 15, "c": 10}}}, {"r": 18, "c": 11, "v": {"mc": {"r": 15, "c": 10}}},
                     {"r": 18, "c": 12, "v": {"mc": {"r": 15, "c": 10}}}, {"r": 19, "c": 0,
                                                                           "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                                 "fs": 11, "fc": "rgb(51, 51, 51)",
                                                                                 "ht": 1, "vt": 1}}, {"r": 19, "c": 1,
                                                                                                      "v": {"bg": None,
                                                                                                            "bl": 0,
                                                                                                            "it": 0,
                                                                                                            "ff": 0,
                                                                                                            "fs": 11,
                                                                                                            "fc": "rgb(51, 51, 51)",
                                                                                                            "ht": 1,
                                                                                                            "vt": 1}},
                     {"r": 19, "c": 2,
                      "v": {"v": "Luckysheet", "ct": {"fa": "General", "t": "g"}, "bg": "rgb(67, 110, 238)", "bl": 0,
                            "it": 0, "ff": 0, "fs": 11, "fc": "rgb(255, 215, 0)", "ht": 1, "vt": 1, "m": "Luckysheet"}},
                     {"r": 19, "c": 3,
                      "v": {"v": "Luckysheet", "ct": {"fa": "General", "t": "g"}, "bg": None, "bl": 1, "it": 0, "ff": 0,
                            "fs": "10", "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1, "m": "Luckysheet"}},
                     {"r": 19, "c": 4,
                      "v": {"v": "Luckysheet", "ct": {"fa": "General", "t": "g"}, "bg": None, "bl": 0, "it": 1, "ff": 0,
                            "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1, "m": "Luckysheet"}}, {"r": 19, "c": 5,
                                                                                                       "v": {
                                                                                                           "v": "Luckysheet",
                                                                                                           "ct": {
                                                                                                               "fa": "General",
                                                                                                               "t": "g"},
                                                                                                           "bg": None,
                                                                                                           "bl": 0,
                                                                                                           "it": 0,
                                                                                                           "ff": 0,
                                                                                                           "fs": 11,
                                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                                           "ht": 1,
                                                                                                           "vt": 1,
                                                                                                           "cl": 1,
                                                                                                           "m": "Luckysheet"}},
                     {"r": 19, "c": 6,
                      "v": {"ct": {"fa": "General", "t": "g"}, "bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11,
                            "fc": "rgb(51, 51, 51)", "ht": "1", "vt": "0", "cl": 1}}, {"r": 19, "c": 7, "v": {
                "ct": {"fa": "General", "t": "g"}, "bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11,
                "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1}}, {"r": 19, "c": 8,
                                                              "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11,
                                                                    "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1}},
                     {"r": 19, "c": 9,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 19, "c": 10,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 20, "c": 0,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 20, "c": 1,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 20, "c": 2,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 20, "c": 3,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 20, "c": 4,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 20, "c": 5,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 20, "c": 6,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 20, "c": 7,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 20, "c": 8,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 20, "c": 9,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 20, "c": 10,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 21, "c": 0,
                                        "v": {"v": "Format", "ct": {"fa": "General", "t": "g"}, "m": "Format",
                                              "bg": None, "bl": 1, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 21, "c": 1,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 21, "c": 2, "v": {
                "ct": {"fa": "##0.00", "t": "n"}, "bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11,
                "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1, "m": "0.25", "v": 0.25}}, {"r": 21, "c": 3, "v": {
                "ct": {"fa": "$#,##0.00_);($#,##0.00)", "t": "n"}, "bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11,
                "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1, "m": "$0.25 ", "v": 0.25}}, {"r": 21, "c": 4, "v": {
                "ct": {"fa": "\"$\" 0.00", "t": "n"}, "bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11,
                "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1, "m": "$ 0.25", "v": 0.25}}, {"r": 21, "c": 5, "v": {
                "ct": {"fa": "0%", "t": "n"}, "bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                "ht": 1, "vt": 1, "m": "25%", "v": 0.25}}, {"r": 21, "c": 6,
                                                            "v": {"ct": {"fa": "# ?/?", "t": "n"}, "bg": None, "bl": 0,
                                                                  "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                                                  "ht": 1, "vt": 1, "m": " 1/4", "v": 0.25}},
                     {"r": 21, "c": 7,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 21, "c": 8,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 21, "c": 9,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 21, "c": 10,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 22, "c": 0,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 22, "c": 1,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 22, "c": 2,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 22, "c": 3,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 22, "c": 4,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 22, "c": 5,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 22, "c": 6,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 22, "c": 7,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 22, "c": 8,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 22, "c": 9,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 22, "c": 10,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 23, "c": 0,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 23, "c": 1,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 23, "c": 2,
                                        "v": {"ct": {"fa": "0.00E+00", "t": "n"}, "bg": None, "bl": 0, "it": 0, "ff": 0,
                                              "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1, "m": "2.50E-01",
                                              "v": 0.25}}, {"r": 23, "c": 3,
                                                            "v": {"v": 0.25, "ct": {"fa": "0.00", "t": "n"},
                                                                  "m": "0.25", "bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                  "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1}},
                     {"r": 23, "c": 4,
                      "v": {"ct": {"fa": "AM/PM h:mm:ss", "t": "d"}, "bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11,
                            "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1, "v": 44032, "m": "AM 12:00:00"}},
                     {"r": 23, "c": 5,
                      "v": {"ct": {"fa": "yyyy/MM/dd", "t": "d"}, "bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11,
                            "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1, "v": 44032, "m": "2020/07/20"}},
                     {"r": 23, "c": 6,
                      "v": {"ct": {"fa": "yyyy\"年\"M\"月\"d\"日\"", "t": "d"}, "bg": None, "bl": 0, "it": 0, "ff": 0,
                            "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1, "v": 44032, "m": "2020年7月20日"}},
                     {"r": 23, "c": 7,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 23, "c": 8,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 23, "c": 9,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 23, "c": 10,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 24, "c": 0,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 24, "c": 1,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 24, "c": 2,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 24, "c": 3,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 24, "c": 4,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 24, "c": 5,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 24, "c": 6,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 24, "c": 7,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 24, "c": 8,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 24, "c": 9,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 24, "c": 10,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 25, "c": 0,
                                                                                     "v": {"v": "Alignment",
                                                                                           "ct": {"fa": "General",
                                                                                                  "t": "g"},
                                                                                           "m": "Alignment", "bg": None,
                                                                                           "bl": 1, "it": 0, "ff": 0,
                                                                                           "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 25, "c": 1,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 25, "c": 2,
                                        "v": {"v": "Top Left", "ct": {"fa": "General", "t": "g"}, "m": "Top Left",
                                              "bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": "1", "vt": "1"}}, {"r": 25, "c": 3, "v": {"v": "Top Center",
                                                                                              "ct": {"fa": "General",
                                                                                                     "t": "g"},
                                                                                              "m": "Top Center",
                                                                                              "bg": None, "bl": 0,
                                                                                              "it": 0, "ff": 0,
                                                                                              "fs": 11,
                                                                                              "fc": "rgb(51, 51, 51)",
                                                                                              "ht": "0", "vt": "1"}},
                     {"r": 25, "c": 4,
                      "v": {"v": "Top Right", "ct": {"fa": "General", "t": "g"}, "m": "Top Right", "bg": None, "bl": 0,
                            "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": "2", "vt": "1"}},
                     {"r": 25, "c": 5,
                      "v": {"v": "Center Left", "ct": {"fa": "General", "t": "g"}, "m": "Center Left", "bg": None,
                            "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": "1", "vt": "0"}},
                     {"r": 25, "c": 6,
                      "v": {"v": "Center Center", "ct": {"fa": "General", "t": "g"}, "m": "Center Center", "bg": None,
                            "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": "0", "vt": "0"}},
                     {"r": 25, "c": 7,
                      "v": {"v": "Center Right", "ct": {"fa": "General", "t": "g"}, "m": "Center Right", "bg": None,
                            "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": "2", "vt": "0"}},
                     {"r": 25, "c": 8,
                      "v": {"v": "Bottom Left", "ct": {"fa": "General", "t": "g"}, "m": "Bottom Left", "bg": None,
                            "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": "1", "vt": "2"}},
                     {"r": 25, "c": 9,
                      "v": {"v": "Bottom Center", "ct": {"fa": "General", "t": "g"}, "m": "Bottom Center", "bg": None,
                            "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": "0", "vt": "2"}},
                     {"r": 25, "c": 10,
                      "v": {"v": "Bottom Right", "ct": {"fa": "General", "t": "g"}, "m": "Bottom Right", "bg": None,
                            "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": "2", "vt": "2"}},
                     {"r": 26, "c": 0,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 26, "c": 1,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 26, "c": 2,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 26, "c": 3,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 26, "c": 4,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 26, "c": 5,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 26, "c": 6,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 26, "c": 7,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 26, "c": 8,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 26, "c": 9,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 26, "c": 10,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 27, "c": 0,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 27, "c": 1,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 27, "c": 2,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 27, "c": 3,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 27, "c": 4,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 27, "c": 5,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 27, "c": 6,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 27, "c": 7,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 27, "c": 8,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 27, "c": 9,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 27, "c": 10,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 28, "c": 0, "v": {"v": "WordWrap",
                                                                                          "ct": {"fa": "General",
                                                                                                 "t": "g"},
                                                                                          "m": "WordWrap", "bg": None,
                                                                                          "bl": 1, "it": 0, "ff": 0,
                                                                                          "fs": 11,
                                                                                          "fc": "rgb(51, 51, 51)",
                                                                                          "ht": 1, "vt": 1}},
                     {"r": 28, "c": 1,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 28, "c": 2,
                                        "v": {"v": "ABCDEFGHIJKLMNOPQRSTUVWXYZ", "ct": {"fa": "General", "t": "g"},
                                              "m": "ABCDEFGHIJKLMNOPQRSTUVWXYZ", "bg": None, "bl": 0, "it": 0, "ff": 0,
                                              "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1, "tb": "2"}},
                     {"r": 28, "c": 3,
                      "v": {"ct": {"fa": "General", "t": "g"}, "bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11,
                            "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1, "tb": "1"}}, {"r": 28, "c": 4,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 28, "c": 5,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 28, "c": 6,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 28, "c": 7,
                                                                   "v": {"bg": None, "bl": 0, "it": 0, "ff": 0,
                                                                         "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                                                                         "vt": 1}}, {"r": 28, "c": 8,
                                                                                     "v": {"bg": None, "bl": 0, "it": 0,
                                                                                           "ff": 0, "fs": 11,
                                                                                           "fc": "rgb(51, 51, 51)",
                                                                                           "ht": 1, "vt": 1}},
                     {"r": 28, "c": 9,
                      "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)", "ht": 1,
                            "vt": 1}}, {"r": 28, "c": 10,
                                        "v": {"bg": None, "bl": 0, "it": 0, "ff": 0, "fs": 11, "fc": "rgb(51, 51, 51)",
                                              "ht": 1, "vt": 1}}, {"r": 29, "c": 0, "v": {
                "ct": {"fa": "General", "t": "inlineStr", "s": [
                    {"ff": "\"times new roman\"", "fc": "rgb(51, 51, 51)", "fs": "12", "cl": 0, "un": 0, "bl": 1,
                     "it": 0, "v": "TextRotate"}]}, "ht": "1", "vt": "0"}}, {"r": 29, "c": 2,
                                                                             "v": {"ct": {"fa": "General", "t": "g"},
                                                                                   "v": "I am Luckysheet text rotate style",
                                                                                   "m": "I am Luckysheet text rotate style",
                                                                                   "tr": "1", "tb": "2", "ht": "1",
                                                                                   "fs": "12"}}, {"r": 29, "c": 3,
                                                                                                  "v": {"ct": {
                                                                                                      "fa": "General",
                                                                                                      "t": "g"},
                                                                                                      "v": "I am Luckysheet text rotate style",
                                                                                                      "m": "I am Luckysheet text rotate style",
                                                                                                      "tr": "2",
                                                                                                      "tb": "2",
                                                                                                      "ht": "0",
                                                                                                      "fs": "12"}},
                     {"r": 29, "c": 4,
                      "v": {"ct": {"fa": "General", "t": "g"}, "v": "I am Luckysheet text rotate style",
                            "m": "I am Luckysheet text rotate style", "tr": "4", "tb": "2", "ht": "1", "fs": "12",
                            "vt": "2"}}, {"r": 29, "c": 5, "v": {"ct": {"fa": "General", "t": "g"},
                                                                 "v": "I am Luckysheet text rotate style",
                                                                 "m": "I am Luckysheet text rotate style", "tr": "5",
                                                                 "tb": "2", "ht": "1", "fs": "12"}}, {"r": 29, "c": 6,
                                                                                                      "v": {"ct": {
                                                                                                          "fa": "General",
                                                                                                          "t": "g"},
                                                                                                          "v": "I am Luckysheet text rotate style",
                                                                                                          "m": "I am Luckysheet text rotate style",
                                                                                                          "tr": "1",
                                                                                                          "tb": "1",
                                                                                                          "ht": "1",
                                                                                                          "fs": "12",
                                                                                                          "vt": "0"}},
                     {"r": 30, "c": 0, "v": {"ct": {"fa": "General", "t": "g"}, "v": "hidden1", "m": "hidden1"}},
                     {"r": 31, "c": 0, "v": {"m": "hidden2", "ct": {"fa": "General", "t": "g"}, "v": "hidden2"}},
                     {"r": 33, "c": 0,
                      "v": {"ct": {"fa": "General", "t": "g"}, "bg": None, "bl": 1, "it": 0, "ff": 0, "fs": 11,
                            "fc": "rgb(51, 51, 51)", "ht": 1, "vt": 1}}],
        "ch_width": 2361,
        "rh_height": 936,
        "luckysheet_select_save": [
            {
                "left": 741,
                "width": 138,
                "top": 796,
                "height": 19,
                "left_move": 741,
                "width_move": 138,
                "top_move": 796,
                "height_move": 19,
                "row": [
                    33,
                    33
                ],
                "column": [
                    6,
                    6
                ],
                "row_focus": 33,
                "column_focus": 6
            }
        ],
        "calcChain": [
            {
                "r": 0,
                "c": 3,
                "index": "0",
                "func": [
                    True,
                    3,
                    "=Formula!A1+Formula!B1"
                ],
                "color": "w",
                "parent": None,
                "chidren": {},
                "times": 0
            }
        ],
        "scrollLeft": 0,
        "scrollTop": 0
    }]

    output_file = "output.xlsx"

    export_luckysheet_to_excel(excel_data, output_file)

    print(f"Excel file has been generated: {output_file}")
